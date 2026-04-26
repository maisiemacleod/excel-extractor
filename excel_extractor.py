"""
Excel Field Extractor v7.2
Drag & drop a folder, configure search parameters, extract cell values from Excel files.
"""
import os
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import threading

try:
    from tkinterdnd2 import DND_FILES, TkinterDnD
    HAS_DND = True
except ImportError:
    HAS_DND = False

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    import xlrd
    HAS_XLRD = True
except ImportError:
    HAS_XLRD = False


# ---------------------------------------------------------------------------
# File discovery
# ---------------------------------------------------------------------------

def find_excel_files(folder):
    results = []
    for root, dirs, files in os.walk(folder):
        for fname in files:
            if fname.lower().endswith(('.xlsx', '.xls')):
                results.append(os.path.join(root, fname))
    return results


# ---------------------------------------------------------------------------
# Sheet selection helpers
# ---------------------------------------------------------------------------

def select_sheet_xlsx(wb, sheet_mode, sheet_val):
    if sheet_mode == 'index':
        n = int(sheet_val)
        idx = (n - 1) if n > 0 else (len(wb.worksheets) + n)
        return wb.worksheets[idx]
    else:
        return wb[str(sheet_val)]


def select_sheet_xls(wb, sheet_mode, sheet_val):
    if sheet_mode == 'index':
        n = int(sheet_val)
        idx = (n - 1) if n > 0 else (wb.nsheets + n)
        return wb.sheet_by_index(idx)
    else:
        return wb.sheet_by_name(str(sheet_val))


# ---------------------------------------------------------------------------
# Cell iteration
# ---------------------------------------------------------------------------

def iter_cells_xlsx_ws(ws):
    for row in ws.iter_rows():
        for cell in row:
            yield cell.row, cell.column, cell.value


def iter_cells_xls_ws(ws):
    for r in range(ws.nrows):
        for c in range(ws.ncols):
            yield r + 1, c + 1, ws.cell_value(r, c)


# ---------------------------------------------------------------------------
# Match logic
# ---------------------------------------------------------------------------

def cell_matches(val, search_text, exact):
    if val is None:
        return False
    s = str(val)
    return s == search_text if exact else search_text in s


# ---------------------------------------------------------------------------
# Core processing
# ---------------------------------------------------------------------------

def process_file(filepath, search_text, exact, occurrence_n,
                 offset_r, offset_d, sheet_mode, sheet_val):
    fname = os.path.basename(filepath)
    ext = os.path.splitext(filepath)[1].lower()

    try:
        if ext == '.xlsx':
            if not HAS_OPENPYXL:
                return {"error": f"{fname}: openpyxl not installed"}
            wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
            try:
                ws = select_sheet_xlsx(wb, sheet_mode, sheet_val)
                cells = list(iter_cells_xlsx_ws(ws))
            except Exception as e:
                wb.close()
                return {"error": f"{fname}: sheet not found ({e})"}
            wb.close()

        elif ext == '.xls':
            if not HAS_XLRD:
                return {"error": f"{fname}: xlrd not installed"}
            wb = xlrd.open_workbook(filepath)
            try:
                ws = select_sheet_xls(wb, sheet_mode, sheet_val)
                cells = list(iter_cells_xls_ws(ws))
            except Exception as e:
                return {"error": f"{fname}: sheet not found ({e})"}

        else:
            return None

    except Exception as e:
        return {"error": f"{fname}: {e}"}

    cells_dict = {(r, c): v for r, c, v in cells}
    matches = [(r, c, v) for r, c, v in cells if cell_matches(v, search_text, exact)]
    total = len(matches)

    if total == 0:
        return None

    try:
        idx = (occurrence_n - 1) if occurrence_n > 0 else (total + occurrence_n)
        match_row, match_col, match_val = matches[idx]
    except IndexError:
        return None

    read_val = cells_dict.get((match_row + offset_d, match_col + offset_r))

    return {
        "filename": fname,
        "field_content": str(match_val),
        "total_occurrences": total,
        "occur_seq": idx + 1,
        "match_row": match_row,
        "match_col": match_col,
        "read_value": str(read_val) if read_val is not None else "",
    }


# ---------------------------------------------------------------------------
# Export
# ---------------------------------------------------------------------------

_HEADERS = [
    "\u6587\u4ef6\u540d",
    "\u5b57\u6bb5\u5185\u5bb9",
    "\u5b57\u6bb5\u51fa\u73b0\u6b21\u6570",
    "\u51fa\u73b0\u7684\u6b21\u5e8f\u6570",
    "\u5339\u914d\u5355\u5143\u683c\u884c",
    "\u5339\u914d\u5355\u5143\u683c\u5217",
    "\u8bfb\u53d6\u5230\u7684\u503c",
]


def _row_values(row):
    return [
        row.get('filename', ''),
        row.get('field_content', ''),
        str(row.get('total_occurrences', '')),
        str(row.get('occur_seq', '')),
        str(row.get('match_row', '')),
        str(row.get('match_col', '')),
        row.get('read_value', ''),
    ]


def export_txt(rows, filepath):
    with open(filepath, 'w', encoding='utf-8') as f:
        f.write('\t'.join(_HEADERS) + '\n')
        for row in rows:
            f.write('\t'.join(_row_values(row)) + '\n')


def export_csv(rows, filepath):
    import csv
    with open(filepath, 'w', encoding='utf-8-sig', newline='') as f:
        writer = csv.writer(f)
        writer.writerow(_HEADERS)
        for row in rows:
            writer.writerow(_row_values(row))


def export_xlsx(rows, filepath):
    if not HAS_OPENPYXL:
        raise RuntimeError("openpyxl not installed")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Results"
    ws.append(_HEADERS)
    for row in rows:
        vals = _row_values(row)
        for i in (2, 3, 4, 5):
            try:
                vals[i] = int(vals[i])
            except (ValueError, TypeError):
                pass
        ws.append(vals)
    wb.save(filepath)


# ---------------------------------------------------------------------------
# Theme — Windows 11 Fluent Light
# ---------------------------------------------------------------------------

BG         = "#f3f3f3"   # window background
SURFACE    = "#ffffff"   # card / panel background
SURFACE2   = "#f9f9f9"   # secondary surface (alternate rows)
ACCENT     = "#0067c0"   # Windows 11 blue
ACCENT_H   = "#1177cc"   # hover blue
ACCENT_D   = "#005ba3"   # pressed/dark blue
TEXT       = "#1a1a1a"   # primary text
SUBTEXT    = "#5a5a5a"   # secondary text
PLACEHOLDER= "#999999"   # hint text
GREEN      = "#107c10"   # success green
GREEN_BG   = "#dff6dd"   # success background
RED        = "#c42b1c"   # error red
RED_BG     = "#fde7e9"   # error row background
BORDER     = "#e0e0e0"   # divider / border
BORDER2    = "#c8c8c8"   # focused border
ROW_ODD    = "#ffffff"
ROW_EVEN   = "#f5f5f5"
DROP_BG    = "#f0f6ff"   # drop zone tint
DROP_BD    = "#b3d0f0"   # drop zone border (blue-tinted)
ENTRY_BG   = "#ffffff"

# Font stack: Segoe UI Variable Text (Win11) → Segoe UI (Win10) → Microsoft YaHei UI (CN)
# tkinter only supports a single font face string — we pick the primary face and rely on
# the OS to fall back for characters not covered (Win10 already has Segoe UI + YaHei).
_FONT_FACE = "Segoe UI Variable Text"
FONT_BODY  = (_FONT_FACE, 10)
FONT_LABEL = (_FONT_FACE, 10)
FONT_BOLD  = (_FONT_FACE, 11, "bold")
FONT_TITLE = (_FONT_FACE, 14, "bold")
FONT_H2    = (_FONT_FACE, 11, "bold")
FONT_SMALL = (_FONT_FACE, 10)
FONT_MONO  = ("Consolas", 10)


def apply_theme(root):
    style = ttk.Style(root)
    style.theme_use('default')

    style.configure(".", background=BG, foreground=TEXT,
                    font=FONT_BODY, borderwidth=0, relief="flat")

    style.configure("TFrame", background=BG)
    style.configure("Surface.TFrame", background=SURFACE)

    # Labels
    style.configure("TLabel", background=BG, foreground=TEXT, font=FONT_LABEL)
    style.configure("Surface.TLabel", background=SURFACE, foreground=TEXT, font=FONT_LABEL)
    style.configure("Sub.TLabel", background=BG, foreground=SUBTEXT, font=FONT_SMALL)
    style.configure("SubSurface.TLabel", background=SURFACE, foreground=SUBTEXT, font=FONT_SMALL)

    # LabelFrame — card style
    style.configure("Card.TLabelframe",
                    background=SURFACE, foreground=TEXT,
                    bordercolor=BORDER, relief="solid", borderwidth=1,
                    padding=12)
    style.configure("Card.TLabelframe.Label",
                    background=SURFACE, foreground=ACCENT,
                    font=FONT_H2)

    # Checkbutton / Radiobutton — use 'clam' theme base so indicators are always visible
    style.configure("TCheckbutton",
                    background=SURFACE, foreground=TEXT,
                    font=FONT_BODY, focuscolor=SURFACE,
                    indicatorcolor=SURFACE, indicatorbackground=SURFACE,
                    indicatorrelief="solid", indicatormargin=3,
                    borderwidth=1)
    style.map("TCheckbutton",
              background=[("active", SURFACE)],
              indicatorcolor=[("selected", ACCENT), ("!selected", SURFACE)],
              indicatorbackground=[("selected", ACCENT), ("!selected", ENTRY_BG)],
              bordercolor=[("focus", ACCENT), ("!focus", BORDER2)])

    style.configure("TRadiobutton",
                    background=SURFACE, foreground=TEXT,
                    font=FONT_BODY, focuscolor=SURFACE,
                    indicatorcolor=SURFACE, indicatorbackground=SURFACE,
                    indicatorrelief="solid", indicatormargin=3,
                    borderwidth=1)
    style.map("TRadiobutton",
              background=[("active", SURFACE)],
              indicatorcolor=[("selected", ACCENT), ("!selected", SURFACE)],
              indicatorbackground=[("selected", ACCENT), ("!selected", ENTRY_BG)],
              bordercolor=[("focus", ACCENT), ("!focus", BORDER2)])

    # Entry
    style.configure("TEntry", fieldbackground=ENTRY_BG, foreground=TEXT,
                    insertcolor=TEXT, borderwidth=1, relief="solid",
                    padding=(7, 5))
    style.map("TEntry",
              bordercolor=[("focus", ACCENT), ("!focus", BORDER)],
              fieldbackground=[("readonly", SURFACE2)])

    # Spinbox
    style.configure("TSpinbox", fieldbackground=ENTRY_BG, foreground=TEXT,
                    background=ENTRY_BG, arrowcolor=SUBTEXT,
                    borderwidth=1, relief="solid", padding=(5, 4))
    style.map("TSpinbox",
              bordercolor=[("focus", ACCENT), ("!focus", BORDER)])

    # Primary button (accent)
    style.configure("Accent.TButton",
                    background=ACCENT, foreground="#ffffff",
                    font=FONT_BOLD, padding=(18, 8),
                    relief="flat", borderwidth=0)
    style.map("Accent.TButton",
              background=[("active", ACCENT_H), ("disabled", "#b0c8e0"), ("pressed", ACCENT_D)],
              foreground=[("disabled", "#e0e0e0")])

    # Secondary button
    style.configure("TButton",
                    background=SURFACE, foreground=ACCENT,
                    font=FONT_BODY, padding=(14, 7),
                    relief="solid", borderwidth=1)
    style.map("TButton",
              background=[("active", "#e8f0f8"), ("pressed", "#d0e4f4")],
              bordercolor=[("active", ACCENT_H), ("!active", BORDER2)],
              foreground=[("disabled", SUBTEXT)])

    # Progressbar
    style.configure("TProgressbar",
                    troughcolor=BORDER, background=ACCENT,
                    borderwidth=0, thickness=4)

    # Treeview
    style.configure("Treeview",
                    background=ROW_ODD, foreground=TEXT,
                    fieldbackground=ROW_ODD, font=FONT_BODY,
                    rowheight=26, borderwidth=0, relief="flat")
    style.configure("Treeview.Heading",
                    background=SURFACE2, foreground=SUBTEXT,
                    font=FONT_BOLD, borderwidth=0, relief="flat",
                    padding=(6, 6))
    style.map("Treeview",
              background=[("selected", "#cce4f7")],
              foreground=[("selected", TEXT)])
    style.map("Treeview.Heading",
              background=[("active", BORDER)])

    # Scrollbar — thin & minimal
    style.configure("TScrollbar",
                    background=SURFACE2, troughcolor=SURFACE2,
                    borderwidth=0, arrowcolor=SUBTEXT,
                    width=8, relief="flat")
    style.map("TScrollbar",
              background=[("active", BORDER2)])

    # Separator
    style.configure("TSeparator", background=BORDER)


# ---------------------------------------------------------------------------
# Rounded-corner frame helper (drawn with Canvas)
# ---------------------------------------------------------------------------

class RoundedCard(tk.Canvas):
    """A canvas that draws a rounded-rectangle card background."""
    RADIUS = 8

    def __init__(self, parent, bg=SURFACE, border=BORDER, **kw):
        super().__init__(parent, bg=parent["bg"], highlightthickness=0, **kw)
        self._card_bg = bg
        self._card_border = border
        self.bind("<Configure>", self._redraw)

    def _redraw(self, _event=None):
        w, h, r = self.winfo_width(), self.winfo_height(), self.RADIUS
        if w < 2 or h < 2:
            return
        self.delete("all")
        x0, y0, x1, y1 = 1, 1, w - 1, h - 1
        pts = [
            x0+r, y0, x1-r, y0,
            x1, y0, x1, y0+r,
            x1, y1-r, x1, y1,
            x1-r, y1, x0+r, y1,
            x0, y1, x0, y1-r,
            x0, y0+r, x0, y0,
        ]
        self.create_polygon(pts, smooth=True,
                            fill=self._card_bg, outline=self._card_border,
                            width=1)


# ---------------------------------------------------------------------------
# GUI
# ---------------------------------------------------------------------------

class App:
    def __init__(self, root):
        self.root = root
        self.root.title("Excel \u5b57\u6bb5\u63d0\u53d6\u5de5\u5177")
        self.root.configure(bg=BG)
        self.root.resizable(True, True)
        self.folder_path = tk.StringVar()
        self.results = []
        apply_theme(root)
        self._build_ui()

    # ------------------------------------------------------------------
    # UI construction
    # ------------------------------------------------------------------

    def _card(self, parent, title, pady=(8, 4)):
        f = ttk.LabelFrame(parent, text=title, style="Card.TLabelframe")
        f.pack(fill='x', padx=16, pady=pady)
        return f

    def _build_ui(self):
        # ---- Title bar ----
        titlebar = tk.Frame(self.root, bg=SURFACE, height=56)
        titlebar.pack(fill='x')
        titlebar.pack_propagate(False)

        # Left: icon + title
        left_title = tk.Frame(titlebar, bg=SURFACE)
        left_title.pack(side='left', padx=16, pady=0)
        tk.Label(left_title, text="\U0001f4ca", bg=SURFACE,
                 font=(_FONT_FACE, 18)).pack(side='left', padx=(0, 8))
        title_col = tk.Frame(left_title, bg=SURFACE)
        title_col.pack(side='left')
        tk.Label(title_col,
                 text="Excel \u5b57\u6bb5\u63d0\u53d6\u5de5\u5177",
                 bg=SURFACE, fg=TEXT, font=FONT_TITLE).pack(anchor='w')
        tk.Label(title_col,
                 text="v7.2  \u2014  \u6279\u91cf\u63d0\u53d6 Excel \u5355\u5143\u683c\u5185\u5bb9",
                 bg=SURFACE, fg=SUBTEXT, font=FONT_SMALL).pack(anchor='w')

        # Bottom border of titlebar
        tk.Frame(self.root, bg=BORDER, height=1).pack(fill='x')

        # ---- Scrollable body ----
        outer = tk.Frame(self.root, bg=BG)
        outer.pack(fill='both', expand=True)

        canvas = tk.Canvas(outer, bg=BG, highlightthickness=0)
        vscroll = ttk.Scrollbar(outer, orient='vertical', command=canvas.yview)
        canvas.configure(yscrollcommand=vscroll.set)
        vscroll.pack(side='right', fill='y')
        canvas.pack(side='left', fill='both', expand=True)

        main = tk.Frame(canvas, bg=BG)
        cw = canvas.create_window((0, 0), window=main, anchor='nw')

        main.bind("<Configure>",
                  lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.bind("<Configure>",
                    lambda e: canvas.itemconfig(cw, width=e.width))
        canvas.bind_all("<MouseWheel>",
                        lambda e: canvas.yview_scroll(int(-1*(e.delta/120)), "units"))

        # ---- Drop zone card ----
        drop_card = self._card(main, "\u8f93\u5165\u6587\u4ef6\u5939", pady=(14, 4))

        self.drop_zone = tk.Frame(drop_card, bg=DROP_BG,
                                   highlightbackground=DROP_BD,
                                   highlightthickness=1,
                                   cursor="hand2")
        self.drop_zone.pack(fill='x', pady=(0, 8))

        dz_inner = tk.Frame(self.drop_zone, bg=DROP_BG)
        dz_inner.pack(pady=18)

        self._drop_icon = tk.Label(dz_inner, text="\U0001f4c2",
                                    bg=DROP_BG, fg=ACCENT,
                                    font=(_FONT_FACE, 26))
        self._drop_icon.pack()

        drop_hint_text = ("\u62d6\u653e\u6587\u4ef6\u5939\u5230\u6b64\u5904"
                          if HAS_DND else
                          "\u70b9\u51fb\u9009\u62e9\u6587\u4ef6\u5939")
        self._drop_hint = tk.Label(dz_inner,
                                    text=drop_hint_text,
                                    bg=DROP_BG, fg=SUBTEXT,
                                    font=FONT_SMALL)
        self._drop_hint.pack(pady=(4, 0))

        self._drop_path = tk.Label(self.drop_zone, text="",
                                    bg=DROP_BG, fg=ACCENT,
                                    font=FONT_BOLD, wraplength=600)
        self._drop_path.pack(pady=(0, 8))

        # Hover effect
        def _dz_enter(_e=None):
            self.drop_zone.config(highlightbackground=ACCENT)
            self._drop_icon.config(fg=ACCENT_H)
        def _dz_leave(_e=None):
            self.drop_zone.config(highlightbackground=DROP_BD)
            self._drop_icon.config(fg=ACCENT)

        for w in (self.drop_zone, dz_inner, self._drop_icon, self._drop_hint):
            w.bind("<Button-1>", self._choose_folder)
            w.bind("<Enter>", _dz_enter)
            w.bind("<Leave>", _dz_leave)

        if HAS_DND:
            try:
                self.drop_zone.drop_target_register(DND_FILES)
                self.drop_zone.dnd_bind('<<Drop>>', self._on_drop)
                self.drop_zone.dnd_bind('<<DragEnter>>', lambda e: e.action)
                self.drop_zone.dnd_bind('<<DragOver>>', lambda e: e.action)
            except Exception:
                pass

        # Browse button
        btn_row = tk.Frame(drop_card, bg=SURFACE)
        btn_row.pack(anchor='w', pady=(0, 2))
        ttk.Button(btn_row, text="\U0001f4c1  \u6d4f\u89c8\u6587\u4ef6\u5939\u2026",
                   command=self._choose_folder).pack(side='left')

        # ---- Parameters card ----
        params_card = self._card(main, "\u67e5\u627e\u53c2\u6570", pady=(8, 4))

        grid = tk.Frame(params_card, bg=SURFACE)
        grid.pack(fill='x')
        grid.columnconfigure(1, weight=1)
        grid.columnconfigure(3, weight=1)

        def lbl(text, row, col):
            tk.Label(grid, text=text, bg=SURFACE, fg=SUBTEXT,
                     font=FONT_LABEL).grid(row=row, column=col, sticky='e',
                                           padx=(0, 8), pady=6)

        lbl("\u5b9a\u4f4d\u5185\u5bb9 X", 0, 0)
        self.search_text = tk.StringVar()
        ttk.Entry(grid, textvariable=self.search_text,
                  width=26).grid(row=0, column=1, sticky='ew', pady=5)

        self.exact_match = tk.BooleanVar(value=False)
        ttk.Checkbutton(grid, text="\u7cbe\u786e\u5339\u914d",
                        variable=self.exact_match,
                        style="TCheckbutton").grid(row=0, column=2,
                                                   padx=(12, 0), pady=5, sticky='w')

        lbl("\u7b2c\u51e0\u6b21\u51fa\u73b0 n", 1, 0)
        self.occurrence_n = tk.IntVar(value=1)
        spin_row = tk.Frame(grid, bg=SURFACE)
        spin_row.grid(row=1, column=1, sticky='w', pady=5)
        ttk.Spinbox(spin_row, from_=-9999, to=9999,
                    textvariable=self.occurrence_n,
                    width=8).pack(side='left')
        tk.Label(spin_row, text="\u8d1f\u6570\u4e3a\u5012\u5e8f",
                 bg=SURFACE, fg=PLACEHOLDER, font=FONT_SMALL).pack(side='left', padx=6)

        # Right side: offsets
        lbl("\u5411\u53f3\u504f\u79fb r", 0, 4)
        self.offset_r = tk.IntVar(value=0)
        ttk.Spinbox(grid, from_=-9999, to=9999, textvariable=self.offset_r,
                    width=8).grid(row=0, column=5, sticky='w', padx=(0, 12), pady=5)

        lbl("\u5411\u4e0b\u504f\u79fb d", 1, 4)
        self.offset_d = tk.IntVar(value=0)
        ttk.Spinbox(grid, from_=-9999, to=9999, textvariable=self.offset_d,
                    width=8).grid(row=1, column=5, sticky='w', padx=(0, 12), pady=5)

        # Separator
        tk.Frame(params_card, bg=BORDER, height=1).pack(fill='x', pady=(8, 6))

        # Sheet selection
        sheet_row = tk.Frame(params_card, bg=SURFACE)
        sheet_row.pack(fill='x')
        tk.Label(sheet_row, text="Sheet \u9009\u62e9",
                 bg=SURFACE, fg=SUBTEXT, font=FONT_LABEL).pack(side='left', padx=(0, 14))

        self.sheet_mode = tk.StringVar(value='index')

        ttk.Radiobutton(sheet_row, text="\u7b2c\u51e0\u4e2a Sheet",
                        variable=self.sheet_mode, value='index',
                        command=self._update_sheet_mode).pack(side='left')
        self.sheet_index = tk.IntVar(value=1)
        self.sheet_index_spin = ttk.Spinbox(sheet_row, from_=-9999, to=9999,
                                             textvariable=self.sheet_index, width=7)
        self.sheet_index_spin.pack(side='left', padx=(4, 16))

        ttk.Radiobutton(sheet_row, text="Sheet \u540d\u79f0",
                        variable=self.sheet_mode, value='name',
                        command=self._update_sheet_mode).pack(side='left')
        self.sheet_name = tk.StringVar()
        self.sheet_name_entry = ttk.Entry(sheet_row, textvariable=self.sheet_name, width=18)
        self.sheet_name_entry.pack(side='left', padx=(4, 0))

        self._update_sheet_mode()

        # ---- Action bar ----
        action_card = self._card(main, "\u64cd\u4f5c", pady=(8, 4))

        action_bar = tk.Frame(action_card, bg=SURFACE)
        action_bar.pack(fill='x')

        # Export format (left side) — parent must be SURFACE for radiobutton bg to match
        fmt_group = tk.Frame(action_bar, bg=SURFACE)
        fmt_group.pack(side='left')
        tk.Label(fmt_group, text="\u5bfc\u51fa\u683c\u5f0f",
                 bg=SURFACE, fg=SUBTEXT, font=FONT_LABEL).pack(side='left', padx=(0, 10))
        self.export_fmt = tk.StringVar(value="xlsx")
        for fmt in [("TXT", "txt"), ("CSV", "csv"), ("XLSX", "xlsx")]:
            ttk.Radiobutton(fmt_group, text=fmt[0],
                            variable=self.export_fmt,
                            value=fmt[1]).pack(side='left', padx=4)

        # Buttons (right side)
        btn_group = tk.Frame(action_bar, bg=SURFACE)
        btn_group.pack(side='right')

        self.export_btn_top = ttk.Button(btn_group, text="\u2193  \u5bfc\u51fa",
                                          command=self._export, state='disabled')
        self.export_btn_top.pack(side='left', padx=(0, 8))

        self.run_btn = ttk.Button(btn_group,
                                   text="\u25b6  \u5f00\u59cb\u63d0\u53d6",
                                   style="Accent.TButton",
                                   command=self._run)
        self.run_btn.pack(side='left')

        # ---- Progress ----
        prog_frame = tk.Frame(main, bg=BG)
        prog_frame.pack(fill='x', padx=16, pady=(0, 4))

        self.progress_var = tk.DoubleVar(value=0)
        self.progress_bar = ttk.Progressbar(prog_frame,
                                             variable=self.progress_var,
                                             maximum=100)
        self.progress_bar.pack(fill='x')
        self.status_label = tk.Label(prog_frame, text="",
                                      bg=BG, fg=SUBTEXT,
                                      font=FONT_SMALL, anchor='w')
        self.status_label.pack(fill='x', pady=(2, 0))

        # ---- Results card ----
        results_card = self._card(main, "\u63d0\u53d6\u7ed3\u679c", pady=(8, 16))

        # Column headers & widths
        cols = ["\u6587\u4ef6\u540d", "\u5b57\u6bb5\u5185\u5bb9",
                "\u51fa\u73b0\u6b21\u6570", "\u6b21\u5e8f\u6570",
                "\u884c", "\u5217", "\u8bfb\u53d6\u5230\u7684\u503c"]
        col_widths = [180, 200, 80, 70, 50, 50, 180]

        tree_wrap = tk.Frame(results_card, bg=BORDER, bd=0)
        tree_wrap.pack(fill='both', expand=True)

        self.tree = ttk.Treeview(tree_wrap, columns=cols,
                                  show='headings', height=12)
        for col, w in zip(cols, col_widths):
            self.tree.heading(col, text=col)
            self.tree.column(col, width=w, minwidth=40, anchor='w')

        self.tree.tag_configure('odd',  background=ROW_ODD)
        self.tree.tag_configure('even', background=ROW_EVEN)
        self.tree.tag_configure('error', background=RED_BG, foreground=RED)

        sb = ttk.Scrollbar(tree_wrap, orient='vertical', command=self.tree.yview)
        self.tree.configure(yscrollcommand=sb.set)
        self.tree.pack(side='left', fill='both', expand=True)
        sb.pack(side='right', fill='y')

        # ---- Bottom summary bar ----
        bottom = tk.Frame(main, bg=BG)
        bottom.pack(fill='x', padx=16, pady=(4, 14))

        self.result_count = tk.Label(bottom, text="",
                                      bg=BG, fg=SUBTEXT, font=FONT_SMALL)
        self.result_count.pack(side='left')

        self.export_btn = ttk.Button(bottom, text="\u2193  \u5bfc\u51fa\u7ed3\u679c",
                                      command=self._export, state='disabled')
        self.export_btn.pack(side='right')

    # ------------------------------------------------------------------
    # Event handlers
    # ------------------------------------------------------------------

    def _update_sheet_mode(self):
        mode = self.sheet_mode.get()
        if mode == 'index':
            self.sheet_index_spin.config(state='normal')
            self.sheet_name_entry.config(state='disabled')
        else:
            self.sheet_index_spin.config(state='disabled')
            self.sheet_name_entry.config(state='normal')

    def _choose_folder(self, event=None):
        path = filedialog.askdirectory(title="\u9009\u62e9\u6587\u4ef6\u5939")
        if path:
            self._set_folder(path)

    def _set_folder(self, path):
        self.folder_path.set(path)
        short = path if len(path) < 65 else "\u2026" + path[-62:]
        self._drop_path.config(text="\u2714  " + short)
        self._drop_hint.config(fg=GREEN)

    def _on_drop(self, event):
        path = event.data.strip()
        if path.startswith('{') and path.endswith('}'):
            path = path[1:-1]
        if os.path.isdir(path):
            self._set_folder(path)
        else:
            messagebox.showwarning(
                "\u63d0\u793a",
                "\u8bf7\u62d6\u653e\u6587\u4ef6\u5939\uff0c\u800c\u975e\u6587\u4ef6")
        return event.action

    def _run(self):
        folder = self.folder_path.get()
        if not folder or not os.path.isdir(folder):
            messagebox.showwarning("\u63d0\u793a", "\u8bf7\u5148\u9009\u62e9\u6587\u4ef6\u5939")
            return
        search_text = self.search_text.get()
        if not search_text:
            messagebox.showwarning("\u63d0\u793a", "\u8bf7\u8f93\u5165\u5b9a\u4f4d\u5185\u5bb9 X")
            return

        sheet_mode = self.sheet_mode.get()
        sheet_val = (self.sheet_index.get() if sheet_mode == 'index'
                     else self.sheet_name.get())
        if sheet_mode == 'name' and not sheet_val:
            messagebox.showwarning("\u63d0\u793a", "\u8bf7\u8f93\u5165 Sheet \u540d\u79f0")
            return

        self.run_btn.config(state='disabled')
        self.export_btn.config(state='disabled')
        self.export_btn_top.config(state='disabled')
        self.results = []
        self.result_count.config(text="")
        for row in self.tree.get_children():
            self.tree.delete(row)
        self.progress_var.set(0)

        def worker():
            files = find_excel_files(folder)
            total = len(files)
            if total == 0:
                self.root.after(0, lambda: self._finish(
                    [], "\u672a\u627e\u5230 Excel \u6587\u4ef6"))
                return
            results = []
            for i, fp in enumerate(files):
                res = process_file(
                    fp, search_text,
                    self.exact_match.get(),
                    self.occurrence_n.get(),
                    self.offset_r.get(),
                    self.offset_d.get(),
                    sheet_mode, sheet_val,
                )
                if res:
                    results.append(res)
                pct = (i + 1) / total * 100
                self.root.after(0, lambda p=pct, f=fp:
                                self._update_progress(p, os.path.basename(f)))
            msg = (f"\u5904\u7406 {total} \u4e2a\u6587\u4ef6\uff0c"
                   f"\u5339\u914d {len(results)} \u6761\u7ed3\u679c")
            self.root.after(0, lambda: self._finish(results, msg))

        threading.Thread(target=worker, daemon=True).start()

    def _update_progress(self, pct, fname):
        self.progress_var.set(pct)
        self.status_label.config(text=f"\u6b63\u5728\u5904\u7406: {fname}")

    def _finish(self, results, msg):
        self.results = results
        self.status_label.config(text=msg)
        self.progress_var.set(100)
        for i, row in enumerate(results):
            tag = ('error' if 'error' in row
                   else 'odd' if i % 2 == 0 else 'even')
            if 'error' in row:
                self.tree.insert('', 'end',
                                  values=(row['error'], '', '', '', '', '', ''),
                                  tags=(tag,))
            else:
                self.tree.insert('', 'end', values=(
                    row['filename'],
                    row['field_content'],
                    row['total_occurrences'],
                    row['occur_seq'],
                    row['match_row'],
                    row['match_col'],
                    row['read_value'],
                ), tags=(tag,))
        self.run_btn.config(state='normal')
        if results:
            self.export_btn.config(state='normal')
            self.export_btn_top.config(state='normal')
            self.result_count.config(
                text=f"\u5171 {len(results)} \u6761\u7ed3\u679c",
                fg=GREEN)

    def _export(self):
        fmt = self.export_fmt.get()
        ftypes = {
            "txt":  [("Text files", "*.txt")],
            "csv":  [("CSV files", "*.csv")],
            "xlsx": [("Excel files", "*.xlsx")],
        }
        path = filedialog.asksaveasfilename(
            defaultextension=f".{fmt}",
            filetypes=ftypes.get(fmt, [("All files", "*.*")]),
            title="\u4fdd\u5b58\u7ed3\u679c")
        if not path:
            return
        try:
            if fmt == 'txt':
                export_txt(self.results, path)
            elif fmt == 'csv':
                export_csv(self.results, path)
            elif fmt == 'xlsx':
                export_xlsx(self.results, path)
            messagebox.showinfo("\u5bfc\u51fa\u6210\u529f",
                                f"\u5df2\u5bfc\u51fa\u5230:\n{path}")
        except Exception as e:
            messagebox.showerror("\u5bfc\u51fa\u5931\u8d25", str(e))


def main():
    if HAS_DND:
        try:
            root = TkinterDnD.Tk()
        except Exception:
            root = tk.Tk()
    else:
        root = tk.Tk()
    root.geometry("900x740")
    root.minsize(660, 540)
    App(root)
    root.mainloop()


if __name__ == "__main__":
    main()
